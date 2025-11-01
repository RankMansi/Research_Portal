"""
Publication Merger and Deduplicator
This script merges publications from Scopus and WoS Excel files, 
removes duplicates, handles dash inflation, and creates unique publication lists.
"""
import pandas as pd
import argparse
import os
import logging
import re
from typing import Dict, List, Set, Tuple
from collections import defaultdict

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def clean_publication_title(title: str) -> str:
    """
    Clean and normalize publication title for comparison.
    Handles various symbols and characters used to inflate publication counts.
    
    Args:
        title: Raw publication title
        
    Returns:
        Cleaned title for deduplication
    """
    if not title or pd.isna(title):
        return ""
    
    # Convert to string and strip whitespace
    title = str(title).strip()
    
    # Skip entries that are just symbols or meaningless content
    if is_inflation_entry(title):
        return ""
    
    # Remove numbering (e.g., "1. ", "2. ", etc.)
    title = re.sub(r'^\d+\.\s*', '', title)
    
    # Remove year in parentheses at the end (e.g., " (2024)")
    title = re.sub(r'\s*\(\d{4}\)\s*$', '', title)
    
    # Remove common inflation patterns
    title = re.sub(r'^[-–—_\s]+$', '', title)  # Only hyphens, dashes, underscores, spaces
    title = re.sub(r'^[•·▪▫◦‣⁃]\s*$', '', title)  # Only bullet points
    title = re.sub(r'^[.,;:!?]+$', '', title)  # Only punctuation
    
    # Convert to lowercase for case-insensitive comparison
    title = title.lower()
    
    # Remove extra whitespace
    title = ' '.join(title.split())
    
    return title

def is_inflation_entry(title: str) -> bool:
    """
    Check if an entry is likely used to inflate publication counts.
    
    Args:
        title: Publication title to check
        
    Returns:
        True if the entry appears to be inflation
    """
    if not title or pd.isna(title):
        return True
    
    title = str(title).strip()
    
    # Check for common inflation patterns
    inflation_patterns = [
        r'^[-–—_\s]+$',  # Only hyphens, dashes, underscores, spaces
        r'^[•·▪▫◦‣⁃]\s*$',  # Only bullet points
        r'^[.,;:!?]+$',  # Only punctuation
        r'^n/a$',  # Not applicable
        r'^na$',  # Not available
        r'^none$',  # None
        r'^null$',  # Null
        r'^undefined$',  # Undefined
        r'^$',  # Empty string
        r'^\s+$',  # Only whitespace
        r'^[-–—]+$',  # Only various dash types
        r'^[_]+$',  # Only underscores
        r'^[.]+$',  # Only dots
        r'^[-–—_.\s]+$',  # Combination of common inflation characters
    ]
    
    for pattern in inflation_patterns:
        if re.match(pattern, title, re.IGNORECASE):
            return True
    
    # Check if title is too short to be meaningful (less than 3 characters)
    if len(title) < 3:
        return True
    
    # Check if title contains only numbers and symbols
    if re.match(r'^[\d\s\-–—_.•·▪▫◦‣⁃.,;:!?]+$', title):
        return True
    
    return False

def extract_publications_from_cell(cell_content: str) -> List[str]:
    """
    Extract individual publications from a cell that may contain multiple publications.
    Filters out inflation entries.
    
    Args:
        cell_content: Cell content that may contain numbered publications
        
    Returns:
        List of individual publication titles
    """
    if not cell_content or pd.isna(cell_content):
        return []
    
    content = str(cell_content).strip()
    
    # Check if the entire cell content is an inflation entry
    if is_inflation_entry(content):
        return []
    
    # Split by newlines to get individual publications
    publications = []
    for line in content.split('\n'):
        line = line.strip()
        if line and not is_inflation_entry(line):
            # Remove numbering if present
            clean_line = re.sub(r'^\d+\.\s*', '', line)
            if clean_line and not is_inflation_entry(clean_line):
                publications.append(clean_line)
    
    return publications

def deduplicate_publications(publications: List[str]) -> List[str]:
    """
    Remove duplicate publications using case-insensitive matching.
    
    Args:
        publications: List of publication titles
        
    Returns:
        List of unique publications
    """
    seen = set()
    unique_publications = []
    
    for pub in publications:
        clean_pub = clean_publication_title(pub)
        if clean_pub and clean_pub not in seen:
            seen.add(clean_pub)
            unique_publications.append(pub)  # Keep original formatting
    
    return unique_publications

def process_excel_file(file_path: str, file_type: str) -> Dict[str, Dict]:
    """
    Process an Excel file and extract author-publication mappings.
    Handles the actual Scopus/WoS Excel structure with columns like:
    Sr.no, Authors, Author full names, Title, Year, Source title, etc.
    
    Args:
        file_path: Path to the Excel file
        file_type: Type of file ("scopus" or "wos")
        
    Returns:
        Dictionary mapping author names to their publications
    """
    logger.info(f"Processing {file_type} file: {file_path}")
    
    try:
        # Read Excel file
        df = pd.read_excel(file_path)
        logger.info(f"Loaded {len(df)} rows from {file_type} file")
        logger.info(f"Available columns: {list(df.columns)}")
        
        # Map columns based on the actual Excel structure
        author_col = None
        title_col = None
        year_col = None
        source_col = None
        
        # Try to identify columns automatically
        for col in df.columns:
            col_lower = str(col).lower().strip()
            logger.info(f"Checking column: '{col}' (lowercase: '{col_lower}')")
            
            # Author columns
            if any(keyword in col_lower for keyword in ['author full names', 'authors', 'author']):
                if 'full names' in col_lower or not author_col:
                    author_col = col
                    logger.info(f"Selected author column: '{col}'")
            
            # Title column - prioritize exact match for "title"
            if col_lower == 'title' or ('title' in col_lower and 'source' not in col_lower):
                title_col = col
                logger.info(f"Selected title column: '{col}'")
            
            # Year column
            if col_lower == 'year' or 'year' in col_lower:
                year_col = col
                logger.info(f"Selected year column: '{col}'")
            
            # Source/Journal column
            if any(keyword in col_lower for keyword in ['source title', 'journal', 'source']):
                source_col = col
                logger.info(f"Selected source column: '{col}'")
        
        if not author_col:
            logger.error(f"Could not find author column in {file_type} file")
            return {}
        
        if not title_col:
            logger.error(f"Could not find title column in {file_type} file")
            return {}
        
        logger.info(f"Using columns - Author: {author_col}, Title: {title_col}, Year: {year_col}, Source: {source_col}")
        
        # Log sample data for debugging
        if len(df) > 0:
            logger.info("Sample data from first row:")
            for col in [author_col, title_col, year_col, source_col]:
                if col:
                    sample_value = str(df.iloc[0][col])[:100] + "..." if len(str(df.iloc[0][col])) > 100 else str(df.iloc[0][col])
                    logger.info(f"  {col}: {sample_value}")
        
        # Process each row and group by author
        author_data = {}
        
        for idx, row in df.iterrows():
            # Extract title first
            title = str(row[title_col]).strip()
            if not title or is_inflation_entry(title):
                continue
            
            # Extract year (optional)
            year = ""
            if year_col and year_col in row:
                year_val = row[year_col]
                if pd.notna(year_val):
                    year = f" ({int(year_val)})" if str(year_val).isdigit() else ""
            
            # Create publication title with only title and year (no source/journal)
            full_title = f"{title}{year}"
            
            # Extract author names - handle multiple authors in one cell
            author_names_str = str(row[author_col]).strip()
            if not author_names_str or author_names_str.lower() in ['author', 'authors', 'author full names', 'nan']:
                continue
            
            # Split multiple authors by semicolon and process each one
            author_names = [name.strip() for name in author_names_str.split(';') if name.strip()]
            
            for author_name in author_names:
                # Clean author name (remove any extra IDs or formatting)
                # Remove patterns like (58920229400) - author IDs
                clean_author = re.sub(r'\s*\(\d+\)\s*$', '', author_name.strip())
                if not clean_author:
                    continue
                
                # Initialize author data if not exists
                if clean_author not in author_data:
                    author_data[clean_author] = {
                        'department': "Unknown",  # Department not available in this format
                        'publications': [],
                        'source': file_type
                    }
                
                # Add publication if not already present (basic deduplication)
                if full_title not in author_data[clean_author]['publications']:
                    author_data[clean_author]['publications'].append(full_title)
        
        # Final deduplication for each author
        for author_name, data in author_data.items():
            data['publications'] = deduplicate_publications(data['publications'])
            logger.info(f"Found {len(data['publications'])} unique publications for {author_name}")
        
        logger.info(f"Processed {len(author_data)} authors from {file_type} file")
        return author_data
        
    except Exception as e:
        logger.error(f"Error processing {file_type} file: {e}")
        return {}

def merge_author_data(scopus_data: Dict, wos_data: Dict) -> Dict[str, Dict]:
    """
    Merge publication data from Scopus and WoS for the same authors.
    
    Args:
        scopus_data: Author data from Scopus file
        wos_data: Author data from WoS file
        
    Returns:
        Merged author data with combined publications
    """
    logger.info("Merging author data from Scopus and WoS files")
    
    merged_data = {}
    all_authors = set(scopus_data.keys()) | set(wos_data.keys())
    
    for author in all_authors:
        scopus_info = scopus_data.get(author, {})
        wos_info = wos_data.get(author, {})
        
        # Combine publications from both sources
        all_publications = []
        if scopus_info.get('publications'):
            all_publications.extend(scopus_info['publications'])
        if wos_info.get('publications'):
            all_publications.extend(wos_info['publications'])
        
        # Deduplicate across sources
        unique_publications = deduplicate_publications(all_publications)
        
        # Use department from either source (prefer non-"Unknown")
        department = "Unknown"
        if scopus_info.get('department') and scopus_info['department'] != "Unknown":
            department = scopus_info['department']
        elif wos_info.get('department') and wos_info['department'] != "Unknown":
            department = wos_info['department']
        
        if unique_publications:
            merged_data[author] = {
                'department': department,
                'publications': unique_publications,
                'scopus_count': len(scopus_info.get('publications', [])),
                'wos_count': len(wos_info.get('publications', [])),
                'total_unique': len(unique_publications)
            }
            logger.info(f"Merged {author}: {len(unique_publications)} unique publications "
                       f"(Scopus: {len(scopus_info.get('publications', []))}, "
                       f"WoS: {len(wos_info.get('publications', []))})")
    
    logger.info(f"Successfully merged data for {len(merged_data)} authors")
    return merged_data

def generate_output_excel(merged_data: Dict, output_path: str) -> Dict:
    """
    Generate the final Excel output with unique publications per author.
    Format: Each professor gets a separate row with their publications in a single cell.
    
    Args:
        merged_data: Merged author publication data
        output_path: Path for the output Excel file
        
    Returns:
        Statistics about the generated data
    """
    logger.info(f"Generating output Excel file: {output_path}")
    
    # Prepare data for Excel - each professor gets their own row
    excel_data = []
    department_stats = defaultdict(lambda: {'authors': 0, 'publications': 0})
    
    for author, data in merged_data.items():
        # Create numbered list of all publications for this professor
        if data['publications']:
            numbered_publications = []
            for i, pub in enumerate(data['publications'], 1):
                numbered_publications.append(f"{i}. {pub}")
            
            # Join all publications with newlines
            all_publications = "\n".join(numbered_publications)
        else:
            all_publications = "No publications found"
        
        # Each professor gets their own row
        excel_data.append({
            'Author': author,
            'Publications': all_publications,
            'Department': data['department'],
            'Total_Publications': data['total_unique']
        })
        
        # Update department statistics
        dept = data['department']
        department_stats[dept]['authors'] += 1
        department_stats[dept]['publications'] += data['total_unique']
    
    # Create DataFrame and save to Excel
    df = pd.DataFrame(excel_data)
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Main data sheet with formatted layout
        df.to_excel(writer, sheet_name="Merged Publications", index=False)
        
        # Get the workbook and worksheet for formatting
        workbook = writer.book
        worksheet = writer.sheets["Merged Publications"]
        
        # Import required classes for formatting
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format header row
        for col_num, column in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
            
            # Set column width
            if column == "Author":
                worksheet.column_dimensions[cell.column_letter].width = 30  # Wider for full names
            elif column == "Publications":
                worksheet.column_dimensions[cell.column_letter].width = 120  # Very wide for full titles
            elif column == "Department":
                worksheet.column_dimensions[cell.column_letter].width = 35
            elif column == "Total_Publications":
                worksheet.column_dimensions[cell.column_letter].width = 18
        
        # Format data rows
        for row_num in range(2, len(df) + 2):
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = border
                
                # Apply different formatting based on column
                column_name = df.columns[col_num - 1]
                
                if column_name == "Author":
                    # Author column - left aligned, bold
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='left', vertical='top')
                elif column_name == "Publications":
                    # Publications column - left aligned with text wrapping
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    # Set row height based on number of publications
                    if cell.value and cell.value != "No publications found":
                        pub_count = cell.value.count('\n') + 1
                        # Estimate row height: base height + (number of publications * line height)
                        estimated_height = max(40, 25 + (pub_count * 18))
                        worksheet.row_dimensions[row_num].height = min(estimated_height, 300)  # Cap at 300
                elif column_name == "Department":
                    # Department column - left aligned
                    cell.alignment = Alignment(horizontal='left', vertical='top')
                elif column_name == "Total_Publications":
                    # Total publications - center aligned
                    cell.alignment = Alignment(horizontal='center', vertical='top')
        
        # Department summary sheet
        dept_summary = []
        for dept, stats in department_stats.items():
            dept_summary.append({
                'Department': dept,
                'Authors': stats['authors'],
                'Total_Publications': stats['publications'],
                'Avg_Publications_Per_Author': round(stats['publications'] / stats['authors'], 2)
            })
        
        dept_df = pd.DataFrame(dept_summary)
        dept_df.to_excel(writer, sheet_name="Department Summary", index=False)
    
    # Calculate overall statistics
    total_authors = len(merged_data)
    total_publications = sum(data['total_unique'] for data in merged_data.values())
    total_departments = len(department_stats)
    
    stats = {
        'total_authors': total_authors,
        'total_publications': total_publications,
        'total_departments': total_departments,
        'avg_publications_per_author': round(total_publications / total_authors, 2) if total_authors > 0 else 0,
        'department_stats': dict(department_stats)
    }
    
    logger.info(f"Generated Excel file with {total_authors} authors, {total_publications} unique publications, {total_departments} departments")
    return stats

def main():
    parser = argparse.ArgumentParser(description="Merge and deduplicate publications from Scopus and WoS Excel files")
    parser.add_argument("--scopus-file", required=True, help="Path to Scopus Excel file")
    parser.add_argument("--wos-file", required=True, help="Path to WoS Excel file")
    parser.add_argument("--output", required=True, help="Path for output Excel file")
    
    args = parser.parse_args()
    
    try:
        # Check if input files exist
        if not os.path.exists(args.scopus_file):
            logger.error(f"Scopus file not found: {args.scopus_file}")
            return
        
        if not os.path.exists(args.wos_file):
            logger.error(f"WoS file not found: {args.wos_file}")
            return
        
        # Process both files
        logger.info("Starting publication merge and deduplication process")
        
        scopus_data = process_excel_file(args.scopus_file, "scopus")
        wos_data = process_excel_file(args.wos_file, "wos")
        
        if not scopus_data and not wos_data:
            logger.error("No data found in either input file")
            return
        
        # Merge the data
        merged_data = merge_author_data(scopus_data, wos_data)
        
        if not merged_data:
            logger.error("No merged data to output")
            return
        
        # Generate output
        stats = generate_output_excel(merged_data, args.output)
        
        # Print summary
        logger.info("✅ Publication merge completed successfully!")
        logger.info(f"📊 Summary:")
        logger.info(f"   - Total authors: {stats['total_authors']}")
        logger.info(f"   - Total unique publications: {stats['total_publications']}")
        logger.info(f"   - Total departments: {stats['total_departments']}")
        logger.info(f"   - Average publications per author: {stats['avg_publications_per_author']}")
        
        logger.info(f"📁 Output saved to: {args.output}")
        
    except Exception as e:
        logger.error(f"Error in main process: {e}")
        raise

if __name__ == "__main__":
    main()
